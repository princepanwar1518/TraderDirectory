"""Backend tests for Lead CRM and shortlisted export (iteration 2)."""
import os
import base64
import pytest
import requests

BASE_URL = os.environ["EXPO_PUBLIC_BACKEND_URL"].rstrip("/")


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


# Shared module state
state = {}


# --- Seed: create or reuse a search to get real traders w/ place_ids ---
def test_seed_search_with_place_ids(session):
    # Try history first
    r = session.get(f"{BASE_URL}/api/history")
    assert r.status_code == 200
    items = r.json()
    chosen = None
    for it in items:
        traders = it.get("traders", [])
        if any(t.get("place_id") for t in traders):
            chosen = it
            break
    if not chosen:
        r2 = session.post(
            f"{BASE_URL}/api/search",
            json={"product": "cement dealers", "location": "Pune"},
            timeout=90,
        )
        assert r2.status_code == 200, r2.text
        chosen = r2.json()
    traders = [t for t in chosen.get("traders", []) if t.get("place_id")]
    assert len(traders) >= 2, "Need at least 2 traders with place_id to test leads"
    state["traders"] = traders
    state["product"] = chosen.get("product")
    state["location"] = chosen.get("location")


# --- PUT /api/leads/{place_id} ---
def test_lead_upsert_creates_with_defaults(session):
    pid = state["traders"][0]["place_id"]
    r = session.put(f"{BASE_URL}/api/leads/{pid}", json={"status": "CONTACTED", "shortlisted": True})
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["place_id"] == pid
    assert d["status"] == "CONTACTED"
    assert d["shortlisted"] is True
    assert "updated_at" in d and len(d["updated_at"]) > 0


def test_lead_upsert_partial_update_preserves_fields(session):
    pid = state["traders"][0]["place_id"]
    # Only update status; shortlisted (True) should be preserved
    r = session.put(f"{BASE_URL}/api/leads/{pid}", json={"status": "QUOTED"})
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["status"] == "QUOTED"
    assert d["shortlisted"] is True


def test_lead_invalid_status_400(session):
    pid = state["traders"][0]["place_id"]
    r = session.put(f"{BASE_URL}/api/leads/{pid}", json={"status": "INVALID_X"})
    assert r.status_code == 400


def test_lead_second_trader_won_not_shortlisted(session):
    pid = state["traders"][1]["place_id"]
    r = session.put(f"{BASE_URL}/api/leads/{pid}", json={"status": "WON", "shortlisted": False})
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["status"] == "WON"
    assert d["shortlisted"] is False


# --- GET /api/leads?place_ids=... ---
def test_get_leads_by_place_ids(session):
    pids = [state["traders"][0]["place_id"], state["traders"][1]["place_id"]]
    q = ",".join(pids)
    r = session.get(f"{BASE_URL}/api/leads", params={"place_ids": q})
    assert r.status_code == 200, r.text
    arr = r.json()
    assert isinstance(arr, list)
    returned_pids = {x["place_id"] for x in arr}
    assert set(pids).issubset(returned_pids)
    # Verify persisted values match prior updates
    by_pid = {x["place_id"]: x for x in arr}
    assert by_pid[pids[0]]["status"] == "QUOTED"
    assert by_pid[pids[0]]["shortlisted"] is True
    assert by_pid[pids[1]]["status"] == "WON"
    assert by_pid[pids[1]]["shortlisted"] is False


# --- POST /api/export-excel only_shortlisted=true ---
def _open_xlsx(b64: str):
    raw = base64.b64decode(b64)
    assert raw[:2] == b"PK"
    import io
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(raw))


def test_export_only_shortlisted_true(session):
    # Defensive: ensure ONLY traders[0] is shortlisted (prior test runs may have
    # shortlisted other traders in this same history).
    for idx, t in enumerate(state["traders"]):
        session.put(
            f"{BASE_URL}/api/leads/{t['place_id']}",
            json={"shortlisted": (idx == 0)},
        )
    payload = {
        "product": state["product"],
        "location": state["location"],
        "traders": state["traders"],
        "only_shortlisted": True,
    }
    r = session.post(f"{BASE_URL}/api/export-excel", json=payload)
    assert r.status_code == 200, r.text
    d = r.json()
    wb = _open_xlsx(d["base64"])
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    assert "Status" in headers and "Shortlisted" in headers
    # Only one trader was shortlisted -> exactly one data row
    data_rows = rows[1:]
    assert len(data_rows) == 1, f"Expected 1 shortlisted row, got {len(data_rows)}"
    status_idx = headers.index("Status")
    short_idx = headers.index("Shortlisted")
    assert data_rows[0][status_idx] == "QUOTED"
    assert data_rows[0][short_idx] == "YES"


def test_export_only_shortlisted_true_none_returns_400(session):
    # Unshortlist all to ensure 400
    for t in state["traders"]:
        session.put(f"{BASE_URL}/api/leads/{t['place_id']}", json={"shortlisted": False})
    payload = {
        "product": state["product"],
        "location": state["location"],
        "traders": state["traders"],
        "only_shortlisted": True,
    }
    r = session.post(f"{BASE_URL}/api/export-excel", json=payload)
    assert r.status_code == 400


def test_export_only_shortlisted_false_includes_status_columns(session):
    payload = {
        "product": state["product"],
        "location": state["location"],
        "traders": state["traders"],
        "only_shortlisted": False,
    }
    r = session.post(f"{BASE_URL}/api/export-excel", json=payload)
    assert r.status_code == 200, r.text
    d = r.json()
    wb = _open_xlsx(d["base64"])
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    assert headers[-2] == "Status" and headers[-1] == "Shortlisted"
    data_rows = rows[1:]
    assert len(data_rows) == len(state["traders"])  # all included
    # Second trader had status WON before unshortlist; status field should persist
    status_idx = headers.index("Status")
    statuses = {row[0]: row[status_idx] for row in data_rows}
    # At least one row should reflect WON for the second trader's name
    second_name = state["traders"][1]["name"]
    assert statuses.get(second_name) == "WON"


def test_export_excel_no_flag_backward_compat(session):
    payload = {
        "product": state["product"],
        "location": state["location"],
        "traders": state["traders"],
    }
    r = session.post(f"{BASE_URL}/api/export-excel", json=payload)
    assert r.status_code == 200
    d = r.json()
    assert d["filename"].endswith(".xlsx")
    assert len(d["base64"]) > 100
