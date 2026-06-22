from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import base64
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone

import serpapi
from openpyxl import Workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# SerpAPI client
SERPAPI_KEY = os.environ.get('SERPAPI_KEY', '')
serp_client = serpapi.Client(api_key=SERPAPI_KEY) if SERPAPI_KEY else None

app = FastAPI()
api_router = APIRouter(prefix="/api")


# ------- Models -------
class Trader(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    address: Optional[str] = None
    phone: Optional[str] = None
    category: Optional[str] = None
    rating: Optional[float] = None
    website: Optional[str] = None
    place_id: Optional[str] = None


class SearchRequest(BaseModel):
    product: str
    location: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None


class SearchResponse(BaseModel):
    id: str
    product: str
    location: str
    count: int
    traders: List[Trader]
    timestamp: str


class ExportRequest(BaseModel):
    product: str
    location: str
    traders: List[Trader]
    only_shortlisted: bool = False


class LeadUpdate(BaseModel):
    status: Optional[str] = None  # NEW | CONTACTED | QUOTED | WON
    shortlisted: Optional[bool] = None


class LeadOut(BaseModel):
    place_id: str
    status: str = "NEW"
    shortlisted: bool = False
    updated_at: str


class ExportResponse(BaseModel):
    filename: str
    base64: str
    mime_type: str


class HistoryItem(BaseModel):
    id: str
    product: str
    location: str
    count: int
    timestamp: str
    traders: List[Trader]


# ------- Helpers -------
def map_serp_to_trader(item: dict) -> Trader:
    name = item.get("title") or ""
    address = item.get("address")
    phone = item.get("phone")
    category = item.get("type")
    if not category:
        types = item.get("types") or []
        category = types[0] if types else None
    rating = item.get("rating")
    website = item.get("website")
    pid = item.get("place_id") or item.get("data_id") or item.get("data_cid")
    return Trader(
        name=name,
        address=address,
        phone=phone,
        category=category,
        rating=float(rating) if rating is not None else None,
        website=website,
        place_id=str(pid) if pid is not None else None,
    )


def perform_serp_search(product: str, location: str,
                        latitude: Optional[float], longitude: Optional[float]) -> List[Trader]:
    if not serp_client:
        raise HTTPException(status_code=500, detail="SerpAPI key not configured")

    params = {
        "engine": "google_maps",
        "type": "search",
        "q": f"{product} in {location}" if location else product,
        "hl": "en",
    }
    if latitude is not None and longitude is not None:
        params["ll"] = f"@{latitude},{longitude},14z"

    try:
        results = serp_client.search(params)
        results_dict = dict(results)
    except Exception as e:
        logger.exception("SerpAPI error")
        raise HTTPException(status_code=502, detail=f"Search provider error: {str(e)}")

    local_results = results_dict.get("local_results") or []
    # local_results may be a dict in some shapes (search engine), handle both
    if isinstance(local_results, dict):
        local_results = local_results.get("places", []) or []

    return [map_serp_to_trader(item) for item in local_results]


# ------- Routes -------
@api_router.get("/")
async def root():
    return {"message": "TraderDirectory API"}


@api_router.post("/search", response_model=SearchResponse)
async def search_traders(req: SearchRequest):
    if not req.product.strip():
        raise HTTPException(status_code=400, detail="Product is required")
    if not req.location.strip() and (req.latitude is None or req.longitude is None):
        raise HTTPException(status_code=400, detail="Location or GPS coordinates required")

    traders = perform_serp_search(req.product, req.location, req.latitude, req.longitude)

    search_id = str(uuid.uuid4())
    ts = datetime.now(timezone.utc).isoformat()

    record = {
        "id": search_id,
        "product": req.product,
        "location": req.location,
        "latitude": req.latitude,
        "longitude": req.longitude,
        "count": len(traders),
        "timestamp": ts,
        "traders": [t.dict() for t in traders],
    }
    await db.searches.insert_one(record)

    return SearchResponse(
        id=search_id,
        product=req.product,
        location=req.location,
        count=len(traders),
        traders=traders,
        timestamp=ts,
    )


@api_router.post("/export-excel", response_model=ExportResponse)
async def export_excel(req: ExportRequest):
    traders_in = req.traders
    # If only_shortlisted, intersect with shortlisted leads
    if req.only_shortlisted:
        place_ids = [t.place_id for t in traders_in if t.place_id]
        leads = await db.leads.find(
            {"place_id": {"$in": place_ids}, "shortlisted": True},
            {"_id": 0},
        ).to_list(1000)
        shortlisted_ids = {ld["place_id"] for ld in leads}
        traders_in = [t for t in traders_in if t.place_id in shortlisted_ids]

    if not traders_in:
        raise HTTPException(status_code=400, detail="No traders to export")

    # Fetch all lead statuses for these traders
    pids = [t.place_id for t in traders_in if t.place_id]
    lead_docs = await db.leads.find({"place_id": {"$in": pids}}, {"_id": 0}).to_list(1000)
    status_by_pid = {ld["place_id"]: ld.get("status", "NEW") for ld in lead_docs}
    shortlisted_by_pid = {ld["place_id"]: ld.get("shortlisted", False) for ld in lead_docs}

    wb = Workbook()
    ws = wb.active
    ws.title = "Traders"

    headers = ["Name", "Category / Industry", "Phone", "Address", "Rating", "Website", "Status", "Shortlisted"]
    ws.append(headers)

    for t in traders_in:
        st = status_by_pid.get(t.place_id or "", "NEW")
        sl = shortlisted_by_pid.get(t.place_id or "", False)
        ws.append([
            t.name or "",
            t.category or "",
            t.phone or "",
            t.address or "",
            t.rating if t.rating is not None else "",
            t.website or "",
            st,
            "YES" if sl else "",
        ])

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
            val = row[0]
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    b64 = base64.b64encode(data).decode("utf-8")

    safe_product = "".join(c for c in req.product if c.isalnum() or c in "-_ ").strip().replace(" ", "_")
    safe_location = "".join(c for c in req.location if c.isalnum() or c in "-_ ").strip().replace(" ", "_")
    filename = f"traders_{safe_product}_{safe_location}.xlsx".lower()

    return ExportResponse(
        filename=filename,
        base64=b64,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@api_router.get("/history", response_model=List[HistoryItem])
async def get_history():
    docs = await db.searches.find({}, {"_id": 0}).sort("timestamp", -1).to_list(50)
    items = []
    for d in docs:
        items.append(HistoryItem(
            id=d.get("id"),
            product=d.get("product", ""),
            location=d.get("location", ""),
            count=d.get("count", 0),
            timestamp=d.get("timestamp", ""),
            traders=[Trader(**t) for t in d.get("traders", [])],
        ))
    return items


@api_router.get("/history/{search_id}", response_model=HistoryItem)
async def get_history_item(search_id: str):
    d = await db.searches.find_one({"id": search_id}, {"_id": 0})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return HistoryItem(
        id=d.get("id"),
        product=d.get("product", ""),
        location=d.get("location", ""),
        count=d.get("count", 0),
        timestamp=d.get("timestamp", ""),
        traders=[Trader(**t) for t in d.get("traders", [])],
    )


@api_router.delete("/history/{search_id}")
async def delete_history_item(search_id: str):
    res = await db.searches.delete_one({"id": search_id})
    return {"deleted": res.deleted_count}


# ------- Lead CRM -------
VALID_STATUSES = {"NEW", "CONTACTED", "QUOTED", "WON"}


@api_router.get("/leads", response_model=List[LeadOut])
async def list_leads(place_ids: Optional[str] = None):
    query = {}
    if place_ids:
        ids = [p.strip() for p in place_ids.split(",") if p.strip()]
        query = {"place_id": {"$in": ids}}
    docs = await db.leads.find(query, {"_id": 0}).to_list(2000)
    return [LeadOut(**d) for d in docs]


@api_router.put("/leads/{place_id}", response_model=LeadOut)
async def upsert_lead(place_id: str, update: LeadUpdate):
    if update.status is not None and update.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of {VALID_STATUSES}")
    now = datetime.now(timezone.utc).isoformat()

    existing = await db.leads.find_one({"place_id": place_id}, {"_id": 0})
    if existing:
        new_status = update.status if update.status is not None else existing.get("status", "NEW")
        new_short = update.shortlisted if update.shortlisted is not None else existing.get("shortlisted", False)
    else:
        new_status = update.status or "NEW"
        new_short = bool(update.shortlisted)

    doc = {
        "place_id": place_id,
        "status": new_status,
        "shortlisted": new_short,
        "updated_at": now,
    }
    await db.leads.update_one({"place_id": place_id}, {"$set": doc}, upsert=True)
    return LeadOut(**doc)


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
